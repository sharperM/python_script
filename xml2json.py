import pandas as pd
import openpyxl

# from PIL import Image

def extract_images_from_xlsx(input_file, output_folder):
    # Load the Excel file
    wb = openpyxl.load_workbook(input_file)

    # Get the first sheet in the workbook
    sheet = wb.active

    # Iterate through rows starting from the second row (index 1)
    for row in sheet.iter_rows(min_row=1, min_col=4,max_col=4, values_only=True):
        cell_value = row[0]

        # Check if the cell contains an image (data type is 'image')
        if isinstance(cell_value, openpyxl.drawing.image.Image):
            # Get the image object from the cell
            image = cell_value.image

            # Save the image to the output folder
            output_filename = f"{output_folder}/image_row{row[0].row}_col{row[0].column}.png"
            image.save(output_filename)
        else:
            print(cell_value)
            print("Cell does not contain an image")


def extract_images_from_excel(input_file, output_folder):
    # Load the Excel file
    wb = openpyxl.load_workbook(input_file)

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print("Processing sheet: {}".format(sheet_name))
        # Iterate through all images in the sheet
        for image in sheet._images:
            img_data = image.image
            print ("Image anchor: {}".format(image.anchor.coord))
            # Save the image to the output folder
            output_filename = f"{output_folder}/{sheet_name}_image_{image.anchor.coord}.png"
            with open(output_filename, "wb") as f:
                f.write(img_data)

def convert_xls_to_json(input_file, output_file):

    jsonname = ['tab1_list','tab1_banner',"tab2_list",'tab3_list']
    for i in range(4):
        # Read the Excel file (assuming the data starts from the second row)
        df = pd.read_excel(input_file, header=None, usecols=[0, 1, 2], skiprows=1,sheet_name=i)

        # Rename the columns to "title" and "desc"
        # 'imageUrl'
        df.columns = ['id','title', 'desc']
        df['imageUrl'] = r"file:///android_asset/" + jsonname[i] + "/"+ jsonname[i] + "_id_" + df['id'].astype(str) + ".webp"
        print(df['imageUrl'])
        json_data = df.to_json(orient='records', force_ascii=False)
    
        # Save JSON data to a file
        with open(jsonname[i]+'.json', 'w',encoding='utf-8') as file:
            file.write(json_data)
        



def readImage():
    from PIL import ImageGrab
    import win32com.client as win32

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel.Workbooks.Open(r'C:\Users\soldi\Desktop\codeup\msp\python_miniconda\nongshidaren.xlsx')

    num = 1
    for sheet in workbook.Worksheets:
        for i, shape in enumerate(sheet.Shapes):
            if shape.Name.startswith('Picture'):
                shape.Copy()
                image = ImageGrab.grabclipboard()   
                image.convert('RGB').save(r'nongshidarenImg\\{}.jpg'.format(num), 'jpeg')
                num+=1
    excel.Quit()

 

if __name__ == "__main__":
    input_file = "nongshidaren.xlsx"   # Replace with the actual path to your xls file
    output_file = "tab3_list.json" # Replace with the desired path for the JSON output
    print("Converting {} to {}".format(input_file, output_file))
    convert_xls_to_json(input_file, output_file)

    # input_file = "path/to/input_file.xlsx"   # Replace with the actual path to your xlsx file
    output_folder = "nongshidarenImg"   # Replace with the desired output folder

